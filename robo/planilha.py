"""
Leitor da planilha de entrada.

A planilha tem o mesmo formato da PREENCHER INFORMACOES do escritorio:
    A: NPJUR
    B: TIPO AGENDAMENTO
    C: TIPO ACOMPANHAMENTO
    D: NOME PETICAO
    E: TIPO PETICAO (PETICIONAMENTOS)
    F: INSTANCIA
    G: ESTADO
    H: CATEGORIA PETICIONAMENTO
    I: TIPO ARQUIVO PETICIONAMENTO
    J: TIPO PETICAO (ESTADOS)
    K: ENVOLVIMENTO
    L: campo de personalizacao, com 2 sentidos dependendo da col C:
       - col C != "PETIÇÃO" (desentranhamento etc.): ENDERECO — o robo edita
         o Word trocando "no endereço constante da inicial" por "no
         endereço a seguir" + endereco + 05 dias pra custas, e poe o
         endereco na Observacao do anexo e do acompanhamento. Vazio = fluxo
         normal.
       - col C == "PETIÇÃO": PEDIDO LIVRE — texto que o robo insere (sem
         negrito) apos "...a fim de expor e requerer o que segue.", usando
         como modelo a peca de col D = "DILAÇÃO DE PRAZO - ATIVAS".
    M: OBSERVACAO PETICAO LIVRE — palavra/texto pra Observacao do anexo e do
       acompanhamento; só usada quando col C == "PETIÇÃO" (col L acima).
"""

from dataclasses import dataclass
from pathlib import Path
import openpyxl


@dataclass
class LinhaProcesso:
    """Uma linha da planilha, ja parseada."""
    linha: int                # numero da linha no Excel (pra mensagens de erro)
    npjur: str
    tipo_agendamento: str
    tipo_acompanhamento: str
    nome_peticao: str
    tipo_peticao_peticionamento: str
    instancia: str
    estado: str
    categoria_peticionamento: str
    tipo_arquivo_peticionamento: str
    tipo_peticao_estados: str
    envolvimento: str
    endereco: str = ""
    obs_peticao_livre: str = ""

    @property
    def eh_peticao_livre(self) -> bool:
        """Col C == 'PETIÇÃO' (sem acento/caixa): fluxo de peticao livre,
        col L passa a ser o pedido (nao o endereco) e col M a observacao."""
        from robo.npjur import _normalizar
        return _normalizar(self.tipo_acompanhamento) == "PETICAO"

    def __str__(self):
        if self.eh_peticao_livre:
            extra = " (PETICAO LIVRE)" if self.endereco else ""
        else:
            extra = " (ENDEREÇO NOVO)" if self.endereco else ""
        return f"NPJUR {self.npjur} [{self.estado}] {self.tipo_acompanhamento}{extra}"


def _texto(v) -> str:
    """Celula -> string. Numeros inteiros viram '1612224', nunca '1612224.0'."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def ler_planilha(caminho: str, aba: str = "PREENCHER INFORMAÇÕES") -> list[LinhaProcesso]:
    """
    Le a planilha e retorna a lista de processos pra processar.

    Pula linhas onde o NPJUR esta vazio.
    """
    caminho_path = Path(caminho)
    if not caminho_path.exists():
        raise FileNotFoundError(
            f"Planilha nao encontrada em: {caminho_path.absolute()}\n"
            f"Coloque sua planilha em '{caminho}' ou ajuste o caminho no config.yaml"
        )

    wb = openpyxl.load_workbook(caminho_path, data_only=True, read_only=True)
    if aba not in wb.sheetnames:
        raise ValueError(
            f"Aba '{aba}' nao encontrada na planilha. "
            f"Abas disponiveis: {wb.sheetnames}"
        )

    sh = wb[aba]
    processos: list[LinhaProcesso] = []

    # Pula o cabecalho (linha 1)
    for idx, row in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue  # linha vazia, pula

        # Normaliza tudo pra string
        valores = [_texto(v) for v in row]
        # Garante 13 colunas (preenche com "" se faltar)
        while len(valores) < 13:
            valores.append("")

        processos.append(LinhaProcesso(
            linha=idx,
            npjur=valores[0],
            tipo_agendamento=valores[1],
            tipo_acompanhamento=valores[2],
            nome_peticao=valores[3],
            tipo_peticao_peticionamento=valores[4],
            instancia=valores[5],
            estado=valores[6],
            categoria_peticionamento=valores[7],
            tipo_arquivo_peticionamento=valores[8],
            tipo_peticao_estados=valores[9],
            envolvimento=valores[10],
            endereco=valores[11],
            obs_peticao_livre=valores[12],
        ))

    wb.close()
    return processos
